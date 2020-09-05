import datetime
import os

import pandas
from openpyxl.styles import *
import openpyxl
from datetime import date

from data.config import projectPath, settings


test_data_path = os.path.join(projectPath, 'test_data//')
test_data_path = test_data_path + settings['test_data_file_name']


def open_excel():
    excel_workbook = openpyxl.load_workbook(test_data_path)
    return excel_workbook


def get_count_of_rows():
    sheet = open_excel("Hotel_list")
    max_row_count = sheet.max_row
    max_column_count = sheet.max_column


def get_test_data(sheet_name, column_name, index):
    workbook = open_excel()
    sheet = workbook[sheet_name]
    max_row_count = sheet.max_row
    max_column_count = sheet.max_column
    for curr_row in range(2, max_row_count + 1):
        for column_count in range(1, max_column_count):
            if str(sheet.cell(1, column_count).value) == str(column_name):
                asked_test_data = str(sheet.cell(index, column_count).value)
                return asked_test_data
    else:
        error_message = 'Required Data NOT available in Data Sheet'
        return error_message


def get_date():
    now = datetime.datetime.now( )
    column_name = "Result_For::" + now.strftime("%d-%B-%Y")
    return column_name


def put_header(sheet_name):
    font = Font(bold='bold')
    path_to_write = os.path.join(projectPath, 'test_data\\')
    path_to_write = path_to_write + settings['file_to_write_results']
    excel_workbook = openpyxl.load_workbook(path_to_write)
    sheet = excel_workbook[sheet_name]
    column_name = get_date( )
    max_column_count = sheet.max_column
    title_counter = max_column_count + 1
    sheet.cell(1, title_counter).value = str(column_name)
    sheet.cell(1, title_counter).font = font
    excel_workbook.save(path_to_write)


def get_max_column_count(sheet_name):
    workbook = open_excel()
    sheet = workbook[sheet_name]
    max_column_count = sheet.max_column
    return max_column_count


def set_results_availability(sheet_name, result_value, index, column_count):
    path_to_write = os.path.join(projectPath, 'test_data\\')
    path_to_write = path_to_write + settings['file_to_write_results']
    excel_workbook = openpyxl.load_workbook(path_to_write)
    sheet = excel_workbook[sheet_name]
    sheet.cell(index, column_count).value = str(result_value)
    excel_workbook.save(path_to_write)


def get_data_dict(sheet_name, current_row_count, status, step_status_dictionary):
    output_dict = {}
    hotel_id = get_test_data(sheet_name, "Hotel_Name", current_row_count)
    start_id = get_test_data(sheet_name, "Start_Date", current_row_count)
    end_date = get_test_data(sheet_name, "End_Date", current_row_count)
    end_date_2 = get_test_data(sheet_name, "End_Date_2", current_row_count)
    end_date_5 = get_test_data(sheet_name, "End_Date_5", current_row_count)
    number_of_passenger = get_test_data(sheet_name, "Number_of_Passenger", current_row_count)
    output_dict.update({"Hotel_Name": hotel_id, "Start_Date": start_id, "End_Date": end_date, "Status_1_Night": step_status_dictionary.get("Status_1_Night"), "End_Date_2": end_date_2, "Status_2_Night": step_status_dictionary.get("Status_2_Night"), "End_Date_5": end_date_5, "Status_5_Night": step_status_dictionary.get("Status_5_Night"), "Number_of_Travellers": number_of_passenger, "Status": status})
    return output_dict


def write_data_to_excel(content):
    """Method to write create and write the data from json file to excel"""
    print("Inside excel")
    today_date = date.today()
    output_excel_file_name = "Booking_Details" + str(today_date)
    path_to_store_excel_file = os.path.join(projectPath, "output_files//")
    output_file_name = path_to_store_excel_file + output_excel_file_name + ".xlsx"
    content_to_add = content
    del content_to_add[0]
    print("Content::", content_to_add)
    pandas.DataFrame(content_to_add).to_excel(output_file_name)
    return output_file_name


# content = [
#    {
#         "test": "test_value"
#     },
#     {
#         "Hotel_Name": "Suite Home Barcelona",
#         "Start_Date": "13-diciembre-2020",
#         "End_Date": "23-diciembre-2020",
#         "Number_of_Travellers": "2 hu\u00e9spedes",
#         "Status": "passed"
#     }
# ]
# write_data_to_excel(content)
